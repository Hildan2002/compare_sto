import pyodbc
from collections import defaultdict
from openpyxl import Workbook

cnxn_str = ("Driver={SQL Server Native Client 11.0};"
            "Server=192.168.10.250;"
            "Database=SBO_NSI_USD_LIVE;"
            "UID=sa;"
            "PWD=P@ssw0rd;")
            
cnxn = pyodbc.connect(cnxn_str)

cursor = cnxn.cursor()
cursor.execute("Select b.itemcode, b.Quantity From OPDN a left join PDN1 b on a.Docentry=b.Docentry WHERE (b.WhsCode = 'WHFG1' or b.WhsCode = 'WHWIPMF1' OR b.WhsCode = 'WHWIPMF2' OR b.WhsCode = 'WHAFSC' OR b.WhsCode = 'WHWIPQC') AND (A.docdate >= '2022-10-01' AND A.DocDate <= '2022-10-31')")  

# Initialisation of defaultdict
output = defaultdict(int)
 
for k, v in cursor:
    output[k] += v

wb = Workbook()

wb.create_sheet("sheet_one")

ws1 = wb['sheet_one']

row_start = 3
col_start = 2


for i in output.items():
    ws1.cell(row=row_start, column=col_start).value = i[0]
    ws1.cell(row=row_start, column=col_start + 1).value = i[1]
    row_start += 1

wb.save('book_eg.xlsx')  