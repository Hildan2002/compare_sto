import pyodbc
from collections import defaultdict
from openpyxl import Workbook, load_workbook
import datetime

dasar_dict = {}
# cnxn_str = ("Driver={SQL Server Native Client 11.0};"
#             "Server=192.168.10.250;"
#             "Database=SBO_NSI_USD_LIVE;"
#             "UID=sa;"
#             "PWD=P@ssw0rd;")
            
# cnxn = pyodbc.connect(cnxn_str)

# def dariProduksi(syalala, tanggalAwal, tanggalAkhir):
#     cursor = cnxn.cursor()
#     cursor.execute(f"SELECT T0.[ItemCode], sum(T0.[CmpltQty]) as 'Receive Qty' FROM OWOR T0 LEFT JOIN WOR1 T1 ON T0.[DocEntry] = T1.[DocEntry] WHERE T0.[CmpltQty] >= 1 and T0.[UserSign] in (19,22) and T0.[Warehouse] = 'WHWIPMF1' and T0.[Status] not in ('C') and T0.[PostDate] >= '{tanggalAwal}' and T0.[PostDate] <= '{tanggalAkhir}' Group by T0.[ItemCode], T0.[CmpltQty]")
#     for k, v in cursor:
#         try:
#             syalala[k]['2OUTPUTPROD'] += v
#         except:
#             syalala[k] = dict.fromkeys(["2OUTPUTPROD"], 0) # defaultdict(lambda:0)
#             syalala[k]['2OUTPUTPROD'] += v
#     return syalala

# def dariSaldoAwal():
#     wb = load_workbook('C:/Users/NB22/Documents/python/STO 31 OKTOBER 2022.xlsx', data_only= True)
#     sheet = wb['SALDOAWAL.']
#     # print(sheet.max_row)
#     for data in range(3, sheet.max_row):
#         dasar_dict[sheet[f'E{data}'].value] = dict.fromkeys(["1SALDOAWAL"], 0) # defaultdict(lambda:0)
#         dasar_dict[sheet[f'E{data}'].value]["1SALDOAWAL"] = sheet[f'F{data}'].value
#     return dasar_dict

# syalala = dariSaldoAwal()
# tanggalAAwal = '2022-10-31'
# tanggalAAkhir = '2022-11-29'
# syalili = dariProduksi(syalala, tanggalAAwal, tanggalAAkhir)
# # dariProduksi(datetime.datetime.strptime(tanggalAAwal, '%Y-%m-%d'), datetime.datetime.strptime(tanggalAAkhir, '%Y-%m-%d'))
# print(dasar_dict)
# # for i in syalili:
# #     print(i)


# # x = input("Masukkan tanggal dengan format (dd-mm-yyyy): ")
# # print(x)

print(dasar_dict)
dasar_dict['partno1'] = {}
print(dasar_dict)
dasar_dict['partno1']['lokasi1'] = 5000
print(dasar_dict)