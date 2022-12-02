from collections import defaultdict
import pyodbc
from openpyxl import Workbook


cnxn_str = ("Driver={SQL Server Native Client 11.0};"
            "Server=192.168.10.250;"
            "Database=SBO_NSI_USD_LIVE;"
            "UID=sa;"
            "PWD=P@ssw0rd;")
            
cnxn = pyodbc.connect(cnxn_str)

cursor = cnxn.cursor()
cursor.execute("select a.ItemCode,sum(a.InQty) as qty from oivl a left join oitm b on a.ItemCode = b.ItemCode where a.LocCode = 'WHSCRAP' and b.ItmsGrpCod in ('110','111','112','113','114') and a.InQty<> 0 and a.CreateDate >= '2022-10-01' and a.CreateDate <= '2022-10-31' group by a.ItemCode")

# Initialisation of defaultdict
output = defaultdict(int)
 
for k, v in cursor:
    setrip = k.rfind('-')
    output[k[0:setrip]] += v

wb = Workbook()

wb.create_sheet("sheet_one")

ws1 = wb['sheet_one']

row_start = 3
col_start = 2

for i in output:
    print(i)

for i in output:
    ws1.cell(row=row_start, column=col_start).value = i
    ws1.cell(row=row_start, column=col_start + 1).value = output[i]
    row_start += 1

wb.save('book_scrap.xlsx')  
