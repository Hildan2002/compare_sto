from collections import defaultdict
import pyodbc
from openpyxl import Workbook


cnxn_str = ("Driver={SQL Server Native Client 11.0};"
            "Server=192.168.10.250;"
            "Database=SBO_NSI_USD_LIVE;"
            "UID=sa;"
            "PWD=P@ssw0rd;")
            
cnxn = pyodbc.connect(cnxn_str)

cursor = cnxn.cursor()
cursor.execute("SELECT T0.[ItemCode], sum(T0.[CmpltQty]) as 'Receive Qty' FROM OWOR T0 LEFT JOIN WOR1 T1 ON T0.[DocEntry] = T1.[DocEntry] WHERE T0.[CmpltQty] >= 1 and T0.[UserSign] in (19,22) and T0.[Warehouse] = 'WHWIPMF1' and T0.[Status]" + 
"not in ('C') and T0.[PostDate] >= '2022-09-30' and T0.[PostDate] <= '2022-10-30' Group by T0.[ItemCode], T0.[CmpltQty]")

# Initialisation of defaultdict
output = defaultdict(int)
 
for k, v in cursor:
    output[k] += v

wb = Workbook()

wb.create_sheet("sheet_one")

ws1 = wb['sheet_one']

row_start = 3
col_start = 2


for i in output.items():
    ws1.cell(row=row_start, column=col_start).value = i[0]
    ws1.cell(row=row_start, column=col_start + 1).value = i[1]
    row_start += 1

wb.save('book_1.xlsx')  
