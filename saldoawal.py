from openpyxl import load_workbook

wb = load_workbook('C:/Users/NB22/Documents/python/STO 31 OKTOBER 2022.xlsx', data_only= True)

sheet = wb['COMP SEPT']

def fungsi():

    return

print(sheet['B5'].value)
print(sheet.max_row-2)
matrixxx = {}

# Access all cells in column A
for data in range(2, sheet.max_row - 2):
    # if data == 5:
    matrixxx[sheet[f'B{data}'].value] = {}
    for kol in ['C','D','E','F']:
        # if kol == 'D':
        matrixxx[sheet[f'B{data}'].value][sheet[f'{kol}1'].value] = sheet[f'{kol}{data}'].value


# matrixxx[sheet[f'B5'].value] = {}
# matrixxx[sheet[f'B5'].value][sheet[f'D1'].value] = sheet[f'D5'].value

print(matrixxx)